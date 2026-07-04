"""Phase 3: ユーザーが手動登録するCSV/Excelデータの取り込み。

未上場の競合データや社内の非公開データなど、IR資料からは取得できない
情報をユーザーが直接登録できるようにする。想定する列: label, kind
(financial/nonfinancial), value, unit, period
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.models import IRDataPoint, IRDataPointKind, IRDataSource

REQUIRED_COLUMNS = {"label", "kind", "value", "unit", "period"}


def _row_to_data_point(row: dict[str, str], document_name: str) -> IRDataPoint:
    value = row.get("value")
    return IRDataPoint(
        label=row["label"],
        kind=IRDataPointKind(row["kind"]),
        value=float(value) if value not in (None, "") else None,
        unit=row.get("unit") or None,
        period=row.get("period") or None,
        source=IRDataSource(document_name=document_name),
    )


def _validate_columns(columns: set[str], path: Path) -> None:
    missing = REQUIRED_COLUMNS - columns
    if missing:
        raise ValueError(f"{path.name}: 必須列が不足しています: {sorted(missing)}")


def load_manual_csv(path: Path) -> list[IRDataPoint]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        _validate_columns(set(reader.fieldnames or []), path)
        return [_row_to_data_point(row, path.name) for row in reader]


def load_manual_excel(path: Path, sheet_name: str | None = None) -> list[IRDataPoint]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = sheet.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows)]
    _validate_columns(set(header), path)

    data_points = []
    for values in rows:
        if all(v is None for v in values):
            continue
        row = dict(zip(header, values))
        row = {k: ("" if v is None else str(v)) for k, v in row.items()}
        data_points.append(_row_to_data_point(row, path.name))
    return data_points


def load_manual_data(path: Path) -> list[IRDataPoint]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_manual_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return load_manual_excel(path)
    raise ValueError(f"未対応のファイル形式です: {suffix}（対応形式: .csv, .xlsx）")
